"""Excel export service — uniform openpyxl helpers.

All ERP exports route through this module so headers, fonts, column sizing
and currency formatting stay consistent across modules. Use `make_workbook`
to start a sheet, append rows with `add_data_row`, then call `to_bytes` to
return the binary payload ready to stream as an attachment.
"""
from __future__ import annotations

import io
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1A365D")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def make_workbook(sheet_title: str = "Export") -> tuple[Workbook, Worksheet]:
    """Return a (workbook, active sheet) pair with the sheet renamed."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Export"
    return wb, ws


def set_headers(ws: Worksheet, headers: Sequence[str]) -> None:
    for idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def add_data_row(ws: Worksheet, row_idx: int, values: Sequence[Any]) -> None:
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = BORDER
        cell.alignment = LEFT
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.number_format = "#,##0"


def autosize_columns(ws: Worksheet, min_width: int = 10, max_width: int = 45) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col:
            text = "" if cell.value is None else str(cell.value)
            if len(text) > longest:
                longest = len(text)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(max_width, longest + 2)
        )


def to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_rows(
    sheet_title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
) -> bytes:
    """One-shot helper: build a workbook from headers + iterable of rows."""
    wb, ws = make_workbook(sheet_title)
    set_headers(ws, headers)
    for offset, row in enumerate(rows, start=2):
        add_data_row(ws, offset, row)
    autosize_columns(ws)
    return to_bytes(wb)
